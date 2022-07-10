import openpyxl,os,re
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill,Alignment,Side,Border

#定义路径
catalogue = "D:/study/visualStudioFile/pythonStudy/resource/excel/"
#遍历目录
filelist = os.listdir(catalogue)

excelNewWb = Workbook()
excelNewWs = excelNewWb.active
#sheet.title = 'song' 设置sheet表名
saveCatalogue = "D:/study/visualStudioFile/pythonStudy/resource2/"
#具体文件,先复制
for file in filelist:
#正则表达式匹配excel文件    
    if re.findall('.xls[x]?$',file):
        filePath = catalogue + file
        excelOldWb = load_workbook(filePath)
        excelOldWs = excelOldWb.active
        for row in excelOldWs.iter_rows(min_row=1,max_row=excelOldWs.max_row,min_col=1,max_col=excelOldWs.max_column,values_only=True):
            excelNewWs.append(row)
        savePath = saveCatalogue + file
        excelNewWb.save(savePath)



#对复制的新文件操作更新样式
filelist = os.listdir(saveCatalogue)
#设置样式
side = Side('thin')#细线条
# 定义表头颜色样式为橙色
header_fill = PatternFill('solid', fgColor='FF7F24')
# 定义表中颜色样式为淡黄色
content_fill = PatternFill('solid', fgColor='FFFFE0')
# 定义表尾颜色样式为淡桔红色
bottom_fill = PatternFill('solid', fgColor='EE9572')

align = Alignment(horizontal='center',vertical='center')
head_border = Border(bottom=side,right=side)
content_border = Border(left=side)

for file in filelist:
    if re.findall('.xls[x]?$',file):
        excelOldWb = load_workbook(filePath)
        excelOldWs = excelOldWb.active
#获取表头
        for cell in excelOldWs[1]:
            cell.fill = header_fill
            cell.alignment = align
            cell.border = head_border
#获取内容
        for row in excelOldWs.iter_rows(min_row=2,max_row=excelOldWs.max_row - 1,min_col=1,max_col=excelOldWs.max_column,values_only=False):
            for cell in row:
                cell.fill = content_fill
                cell.alignment = align
                cell.border = content_border
                print(cell.value)
#获取表尾
        for cell in excelOldWs[excelOldWs.max_row]:
            cell.fill = bottom_fill
            cell.alignment = align
        savePath = saveCatalogue + file
        excelNewWb.save(savePath)

